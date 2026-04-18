from fastapi import APIRouter, Depends, HTTPException, Query, status
from sqlalchemy.orm import Session
from sqlalchemy import and_, func
from typing import Optional, List
from datetime import date, datetime
from decimal import Decimal

from app.core.database import get_db
from app.core.security import get_current_user, require_admin
from app.models.user import User, Case, MonthlyAidCycle, MonthlyAidTransaction
from app.schemas.monthly_aid import (
    MonthlyCycleResponse, MonthlyTransactionResponse,
    MonthlyTransactionUpdate, MonthlyTransactionListResponse,
    MonthlyDashboardStats, TransactionStatus, CycleStatus
)
from app.services.audit_service import create_audit_log

from fastapi.responses import StreamingResponse
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from urllib.parse import quote


router = APIRouter(tags=["Monthly Aid"])

# ============ Helper Functions ============

def get_or_create_current_cycle(db: Session) -> MonthlyAidCycle:
    """Get current month cycle or create it automatically"""
    current_year = date.today().year
    current_month = date.today().month
    
    cycle = db.query(MonthlyAidCycle).filter(
        MonthlyAidCycle.year == current_year,
        MonthlyAidCycle.month == current_month,
        MonthlyAidCycle.is_deleted == False
    ).first()
    
    if not cycle:
        cycle = MonthlyAidCycle(
            year=current_year,
            month=current_month,
            status=CycleStatus.OPEN  # Auto-open when created
        )
        db.add(cycle)
        db.commit()
        db.refresh(cycle)
    
    return cycle

def get_current_cycle(db: Session) -> Optional[MonthlyAidCycle]:
    """Get current month cycle without creating it"""
    current_year = date.today().year
    current_month = date.today().month
    
    return db.query(MonthlyAidCycle).filter(
        MonthlyAidCycle.year == current_year,
        MonthlyAidCycle.month == current_month,
        MonthlyAidCycle.is_deleted == False
    ).first()

# ============ Cycle Management ============

@router.get("/monthly/cycles/current", response_model=MonthlyCycleResponse)
def get_current_cycle_endpoint(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get or create current month cycle"""
    cycle = get_or_create_current_cycle(db)
    
    # Get stats
    transactions = db.query(MonthlyAidTransaction).filter(
        MonthlyAidTransaction.monthly_aid_cycle_id == cycle.id,
        MonthlyAidTransaction.is_deleted == False
    ).all()
    
    total_amount = sum(t.monthly_amount for t in transactions)
    delivered_count = len([t for t in transactions if t.status == TransactionStatus.DELIVERED])
    pending_count = len([t for t in transactions if t.status == TransactionStatus.PENDING])
    missed_count = len([t for t in transactions if t.status == TransactionStatus.MISSED])
    
    return {
        "id": cycle.id,
        "year": cycle.year,
        "month": cycle.month,
        "status": cycle.status,
        "created_at": cycle.created_at,
        "updated_at": cycle.updated_at,
        "total_transactions": len(transactions),
        "total_amount": total_amount,
        "delivered_count": delivered_count,
        "pending_count": pending_count,
        "missed_count": missed_count
    }

@router.post("/monthly/cycles/current/generate")
def generate_current_cycle_transactions(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Generate transactions for current month cycle"""
    cycle = get_or_create_current_cycle(db)
    
    if cycle.status != CycleStatus.OPEN:
        raise HTTPException(status_code=400, detail="Cycle must be open to generate transactions")
    
    # Get all cases with monthly aid enabled
    monthly_cases = db.query(Case).filter(
        Case.is_monthly_aid == True,
        Case.is_deleted == False,
        Case.monthly_aid_amount > 0
    ).all()
    
    created_count = 0
    for case in monthly_cases:
        # Check if transaction already exists
        existing = db.query(MonthlyAidTransaction).filter(
            MonthlyAidTransaction.monthly_aid_cycle_id == cycle.id,
            MonthlyAidTransaction.case_id == case.id,
            MonthlyAidTransaction.is_deleted == False
        ).first()
        
        if not existing:
            transaction = MonthlyAidTransaction(
                monthly_aid_cycle_id=cycle.id,
                case_id=case.id,
                monthly_amount=case.monthly_aid_amount,
                status=TransactionStatus.PENDING
            )
            db.add(transaction)
            created_count += 1
    
    db.commit()
    
    create_audit_log(
        db=db,
        user_id=current_user.id,
        action_type="GENERATE",
        entity_name="MonthlyAidTransaction",
        entity_id=cycle.id,
        description=f"Generated {created_count} transactions for {cycle.year}-{cycle.month}",
        ip_address="127.0.0.1"
    )
    
    return {"message": f"Generated {created_count} transactions successfully", "count": created_count}

@router.post("/monthly/cycles/current/close")
def close_current_cycle(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Close current month cycle"""
    cycle = get_current_cycle(db)
    
    if not cycle:
        raise HTTPException(status_code=404, detail="No cycle found for current month")
    
    cycle.status = CycleStatus.CLOSED
    db.commit()
    
    create_audit_log(
        db=db,
        user_id=current_user.id,
        action_type="CLOSE",
        entity_name="MonthlyAidCycle",
        entity_id=cycle.id,
        description=f"Closed cycle {cycle.year}-{cycle.month}",
        ip_address="127.0.0.1"
    )
    
    return {"message": "Cycle closed successfully"}

@router.get("/monthly/cycles/history")
def get_cycle_history(
    page: int = Query(1, ge=1),
    size: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get past cycles history (for reports)"""
    current_year = date.today().year
    current_month = date.today().month
    
    query = db.query(MonthlyAidCycle).filter(
        MonthlyAidCycle.is_deleted == False,
        ~((MonthlyAidCycle.year == current_year) & (MonthlyAidCycle.month == current_month))
    ).order_by(MonthlyAidCycle.year.desc(), MonthlyAidCycle.month.desc())
    
    total = query.count()
    items = query.offset((page - 1) * size).limit(size).all()
    
    result_items = []
    for cycle in items:
        transactions = db.query(MonthlyAidTransaction).filter(
            MonthlyAidTransaction.monthly_aid_cycle_id == cycle.id,
            MonthlyAidTransaction.is_deleted == False
        ).all()
        
        total_amount = sum(t.monthly_amount for t in transactions)
        delivered_count = len([t for t in transactions if t.status == TransactionStatus.DELIVERED])
        
        result_items.append({
            "id": cycle.id,
            "year": cycle.year,
            "month": cycle.month,
            "status": cycle.status,
            "total_transactions": len(transactions),
            "total_amount": total_amount,
            "delivered_count": delivered_count,
            "created_at": cycle.created_at
        })
    
    return {
        "items": result_items,
        "total": total,
        "page": page,
        "size": size,
        "pages": (total + size - 1) // size
    }

# ============ Delivery Management ============

@router.get("/monthly/current/transactions", response_model=MonthlyTransactionListResponse)
def get_current_transactions(
    page: int = Query(1, ge=1),
    size: int = Query(10, ge=1, le=100),
    status: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get transactions for current month cycle"""
    cycle = get_or_create_current_cycle(db)
    
    query = db.query(MonthlyAidTransaction).filter(
        MonthlyAidTransaction.monthly_aid_cycle_id == cycle.id,
        MonthlyAidTransaction.is_deleted == False
    )
    
    if status:
        query = query.filter(MonthlyAidTransaction.status == status)
    
    if search:
        query = query.join(Case).filter(
            Case.full_name.ilike(f"%{search}%")
        )
    
    total = query.count()
    items = query.order_by(MonthlyAidTransaction.id).offset((page - 1) * size).limit(size).all()
    
    result_items = []
    for t in items:
        case = db.query(Case).filter(Case.id == t.case_id).first()
        delivered_by = db.query(User).filter(User.id == t.delivered_by_user_id).first() if t.delivered_by_user_id else None
        
        result_items.append({
            "id": t.id,
            "monthly_aid_cycle_id": t.monthly_aid_cycle_id,
            "case_id": t.case_id,
            "monthly_amount": float(t.monthly_amount),
            "status": t.status,
            "delivered_date": t.delivered_date.isoformat() if t.delivered_date else None,
            "delivered_by_user_id": t.delivered_by_user_id,
            "delivered_by_name": delivered_by.full_name if delivered_by else None,
            "notes": t.notes,
            "case_name": case.full_name if case else None,
            "case_file_number": case.file_number if case else None,
            "case_phone": case.phone_number_1 if case else None,
            "created_at": t.created_at,
            "updated_at": t.updated_at
        })
    
    return {
        "items": result_items,
        "total": total,
        "page": page,
        "size": size,
        "pages": (total + size - 1) // size
    }

@router.put("/monthly/transactions/{transaction_id}/deliver")
def mark_delivered(
    transaction_id: int,
    transaction_data: MonthlyTransactionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    transaction = db.query(MonthlyAidTransaction).filter(
        MonthlyAidTransaction.id == transaction_id,
        MonthlyAidTransaction.is_deleted == False
    ).first()
    
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    transaction.status = TransactionStatus.DELIVERED
    transaction.delivered_date = date.today()
    transaction.delivered_by_user_id = current_user.id
    if transaction_data.notes:
        transaction.notes = transaction_data.notes
    
    db.commit()
    
    return {"message": "Transaction marked as delivered"}

@router.put("/monthly/transactions/bulk-deliver")
def bulk_mark_delivered(
    transaction_ids: List[int],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    transactions = db.query(MonthlyAidTransaction).filter(
        MonthlyAidTransaction.id.in_(transaction_ids),
        MonthlyAidTransaction.is_deleted == False
    ).all()
    
    for transaction in transactions:
        transaction.status = TransactionStatus.DELIVERED
        transaction.delivered_date = date.today()
        transaction.delivered_by_user_id = current_user.id
    
    db.commit()
    
    return {"message": f"{len(transactions)} transactions marked as delivered"}

# ============ Dashboard ============

@router.get("/monthly/dashboard", response_model=MonthlyDashboardStats)
def get_dashboard_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Total monthly aid cases
    total_monthly_cases = db.query(Case).filter(
        Case.is_monthly_aid == True,
        Case.is_deleted == False
    ).count()
    
    # Current cycle
    current_cycle = get_current_cycle(db)
    
    stats = {
        "total_monthly_cases": total_monthly_cases,
        "current_cycle_id": None,
        "current_cycle_name": None,
        "current_cycle_status": None,
        "current_cycle_total_transactions": 0,
        "current_cycle_total_amount": 0,
        "current_cycle_delivered_count": 0,
        "current_cycle_pending_count": 0,
        "current_cycle_missed_count": 0,
        "current_cycle_completion_percentage": 0
    }
    
    if current_cycle:
        transactions = db.query(MonthlyAidTransaction).filter(
            MonthlyAidTransaction.monthly_aid_cycle_id == current_cycle.id,
            MonthlyAidTransaction.is_deleted == False
        ).all()
        
        total_amount = sum(float(t.monthly_amount) for t in transactions)
        delivered_count = len([t for t in transactions if t.status == TransactionStatus.DELIVERED])
        pending_count = len([t for t in transactions if t.status == TransactionStatus.PENDING])
        missed_count = len([t for t in transactions if t.status == TransactionStatus.MISSED])
        
        completion = (delivered_count / len(transactions) * 100) if transactions else 0
        
        stats.update({
            "current_cycle_id": current_cycle.id,
            "current_cycle_name": f"{current_cycle.year}-{current_cycle.month:02d}",
            "current_cycle_status": current_cycle.status,
            "current_cycle_total_transactions": len(transactions),
            "current_cycle_total_amount": total_amount,
            "current_cycle_delivered_count": delivered_count,
            "current_cycle_pending_count": pending_count,
            "current_cycle_missed_count": missed_count,
            "current_cycle_completion_percentage": round(completion, 2)
        })
    
    return stats

@router.post("/monthly/cycles/current/open")
def open_current_cycle(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Open current month cycle"""
    cycle = get_current_cycle(db)
    
    if not cycle:
        raise HTTPException(status_code=404, detail="No cycle found for current month")
    
    cycle.status = CycleStatus.OPEN
    db.commit()
    
    create_audit_log(
        db=db,
        user_id=current_user.id,
        action_type="OPEN",
        entity_name="MonthlyAidCycle",
        entity_id=cycle.id,
        description=f"Opened cycle {cycle.year}-{cycle.month}",
        ip_address="127.0.0.1"
    )
    
    return {"message": "Cycle opened successfully"}




@router.get("/monthly/cycles/{cycle_id}/details")
def get_cycle_details(
    cycle_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get detailed information about a specific cycle"""
    cycle = db.query(MonthlyAidCycle).filter(
        MonthlyAidCycle.id == cycle_id,
        MonthlyAidCycle.is_deleted == False
    ).first()
    
    if not cycle:
        raise HTTPException(status_code=404, detail="Cycle not found")
    
    # Get all transactions for this cycle
    transactions = db.query(MonthlyAidTransaction).filter(
        MonthlyAidTransaction.monthly_aid_cycle_id == cycle_id,
        MonthlyAidTransaction.is_deleted == False
    ).all()
    
    # Get case details for each transaction
    transaction_details = []
    for t in transactions:
        case = db.query(Case).filter(Case.id == t.case_id).first()
        delivered_by = db.query(User).filter(User.id == t.delivered_by_user_id).first() if t.delivered_by_user_id else None
        
        transaction_details.append({
            "id": t.id,
            "case_id": t.case_id,
            "case_name": case.full_name if case else None,
            "case_file_number": case.file_number if case else None,
            "case_phone": case.phone_number_1 if case else None,
            "monthly_amount": float(t.monthly_amount),
            "status": t.status,
            "delivered_date": t.delivered_date.isoformat() if t.delivered_date else None,
            "delivered_by": delivered_by.full_name if delivered_by else None,
            "notes": t.notes
        })
    
    # Calculate summary statistics
    total_amount = sum(float(t.monthly_amount) for t in transactions)
    delivered_count = len([t for t in transactions if t.status == TransactionStatus.DELIVERED])
    pending_count = len([t for t in transactions if t.status == TransactionStatus.PENDING])
    missed_count = len([t for t in transactions if t.status == TransactionStatus.MISSED])
    completion_percentage = (delivered_count / len(transactions) * 100) if transactions else 0
    
    return {
        "cycle": {
            "id": cycle.id,
            "year": cycle.year,
            "month": cycle.month,
            "status": cycle.status,
            "created_at": cycle.created_at,
            "updated_at": cycle.updated_at
        },
        "summary": {
            "total_transactions": len(transactions),
            "total_amount": total_amount,
            "delivered_count": delivered_count,
            "pending_count": pending_count,
            "missed_count": missed_count,
            "completion_percentage": round(completion_percentage, 2)
        },
        "transactions": transaction_details
    }
    
    
    
from fastapi.responses import StreamingResponse
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from urllib.parse import quote

@router.get("/monthly/cycles/{cycle_id}/export")
def export_cycle_to_excel(
    cycle_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export cycle transactions to Excel"""
    try:
        cycle = db.query(MonthlyAidCycle).filter(
            MonthlyAidCycle.id == cycle_id,
            MonthlyAidCycle.is_deleted == False
        ).first()
        
        if not cycle:
            raise HTTPException(status_code=404, detail="Cycle not found")
        
        # Get all transactions
        transactions = db.query(MonthlyAidTransaction).filter(
            MonthlyAidTransaction.monthly_aid_cycle_id == cycle_id,
            MonthlyAidTransaction.is_deleted == False
        ).all()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Cycle_{cycle.year}_{cycle.month}"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title row
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f'Monthly Cycle Report - {cycle.year}-{cycle.month}'
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Summary rows (use English for compatibility)
        ws['A3'] = 'Month:'
        ws['B3'] = f'{cycle.year}-{cycle.month}'
        ws['A4'] = 'Total Transactions:'
        ws['B4'] = len(transactions)
        
        total_amount = sum(float(t.monthly_amount) for t in transactions)
        ws['A5'] = 'Total Amount:'
        ws['B5'] = f'{total_amount:.2f} EGP'
        
        delivered_count = len([t for t in transactions if t.status == TransactionStatus.DELIVERED])
        ws['A6'] = 'Delivered:'
        ws['B6'] = delivered_count
        
        pending_count = len([t for t in transactions if t.status == TransactionStatus.PENDING])
        ws['A7'] = 'Pending:'
        ws['B7'] = pending_count
        
        missed_count = len([t for t in transactions if t.status == TransactionStatus.MISSED])
        ws['A8'] = 'Missed:'
        ws['B8'] = missed_count
        
        # Headers row (use English for compatibility)
        headers = ['#', 'File Number', 'Beneficiary Name', 'Phone', 'Amount', 'Status', 'Delivery Date', 'Delivered By']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        for idx, t in enumerate(transactions, 1):
            case = db.query(Case).filter(Case.id == t.case_id).first()
            delivered_by = db.query(User).filter(User.id == t.delivered_by_user_id).first() if t.delivered_by_user_id else None
            
            row = 10 + idx
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=2, value=case.file_number if case else '').border = border
            ws.cell(row=row, column=3, value=case.full_name if case else '').border = border
            ws.cell(row=row, column=4, value=case.phone_number_1 if case else '').border = border
            ws.cell(row=row, column=5, value=float(t.monthly_amount)).border = border
            status_text = {'Delivered': 'Delivered', 'Pending': 'Pending', 'Missed': 'Missed'}.get(t.status, t.status)
            ws.cell(row=row, column=6, value=status_text).border = border
            delivered_date_str = t.delivered_date.isoformat() if t.delivered_date else ''
            ws.cell(row=row, column=7, value=delivered_date_str).border = border
            ws.cell(row=row, column=8, value=delivered_by.full_name if delivered_by else '').border = border
        
        # Auto-adjust column widths
        for col_idx in range(1, 9):
            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create filename with proper encoding (use English only for filename)
        filename = f"cycle_{cycle.year}_{cycle.month}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
        )
    except Exception as e:
        print(f"Excel export error: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")



@router.post("/monthly/cycles/create-test")
def create_test_cycle(
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Create a test cycle for specific month (Admin only - for testing)"""
    
    # Check if cycle already exists
    existing = db.query(MonthlyAidCycle).filter(
        MonthlyAidCycle.year == year,
        MonthlyAidCycle.month == month,
        MonthlyAidCycle.is_deleted == False
    ).first()
    
    if existing:
        return {"message": f"Cycle for {year}-{month} already exists", "cycle_id": existing.id}
    
    cycle = MonthlyAidCycle(
        year=year,
        month=month,
        status=CycleStatus.OPEN
    )
    db.add(cycle)
    db.commit()
    db.refresh(cycle)
    
    
    # Generate transactions for this cycle
    monthly_cases = db.query(Case).filter(
        Case.is_monthly_aid == True,
        Case.is_deleted == False,
        Case.monthly_aid_amount > 0
    ).all()
    
    created_count = 0
    for case in monthly_cases:
        transaction = MonthlyAidTransaction(
            monthly_aid_cycle_id=cycle.id,
            case_id=case.id,
            monthly_amount=case.monthly_aid_amount,
            status=TransactionStatus.PENDING
        )
        db.add(transaction)
        created_count += 1
    
    db.commit()
    
    return {
        "message": f"Created cycle for {year}-{month} with {created_count} transactions",
        "cycle_id": cycle.id
    }